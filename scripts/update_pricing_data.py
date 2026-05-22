from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SOURCE_FILE = ROOT / "Precificacao_Servicos_JM_Edificacoesesse.xlsx"
OUTPUT_FILE = ROOT / "pricing-data.js"


def normalize_space(value: object | None) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_price(value: object | None) -> object | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else round(value, 2)
    return normalize_space(value)


def slugify(value: str) -> str:
    lowered = value.lower()
    lowered = re.sub(r"[^\w\s-]", "", lowered, flags=re.UNICODE)
    lowered = re.sub(r"[-\s]+", "-", lowered, flags=re.UNICODE)
    return lowered.strip("-")


def clean_group_title(raw_text: str) -> str:
    text = normalize_space(raw_text)
    text = re.sub(r"^\d+\.\s*", "", text)
    return text


def extract_unit(description: str) -> tuple[str, str | None]:
    match = re.search(r"\[(.*?)\]\s*$", description)
    if not match:
        return description, None

    unit = normalize_space(match.group(1))
    cleaned = re.sub(r"\s*\[(.*?)\]\s*$", "", description).strip()
    return cleaned, unit or None


def build_item(
    *,
    sheet_name: str,
    group_title: str | None,
    group_note: str | None,
    row_index: int,
    item_id: int,
    description: str,
    unit: str | None,
    reference_price: object | None,
    jm_price: object | None,
) -> dict[str, object | None]:
    price_mode = "fixed"
    if isinstance(jm_price, str):
        lowered = jm_price.lower()
        if "combinar" in lowered:
            price_mode = "consult"
        else:
            price_mode = "formula"

    comparison = "none"
    if isinstance(reference_price, (int, float)) and isinstance(jm_price, (int, float)):
        if jm_price < reference_price:
            comparison = "below"
        elif jm_price > reference_price:
            comparison = "above"
        else:
            comparison = "equal"

    return {
        "uid": f"{slugify(sheet_name)}-{row_index}-{item_id}",
        "sheet": sheet_name,
        "sheetSlug": slugify(sheet_name),
        "group": group_title,
        "groupSlug": slugify(group_title or sheet_name),
        "groupNote": group_note,
        "rowIndex": row_index,
        "itemNumber": item_id,
        "description": description,
        "unit": unit,
        "referencePrice": reference_price,
        "jmPrice": jm_price,
        "priceMode": price_mode,
        "comparison": comparison,
        "searchIndex": " ".join(
            filter(
                None,
                [
                    sheet_name,
                    group_title or "",
                    group_note or "",
                    description,
                    unit or "",
                    str(jm_price or ""),
                    str(reference_price or ""),
                ],
            )
        ),
    }


def parse_workbook() -> dict[str, object]:
    workbook = load_workbook(SOURCE_FILE, data_only=True)
    items: list[dict[str, object | None]] = []
    groups: list[dict[str, object | None]] = []

    for worksheet in workbook.worksheets:
        current_group: str | None = None
        current_note: str | None = None
        is_projects_sheet = worksheet.title == "Projetos e Engenharia"

        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            a, b, c, d = (list(row) + [None, None, None, None])[:4]

            if isinstance(a, str) and b is None and c is None and d is None:
                text = normalize_space(a)

                if not text or text.startswith("TABELA DE") or "JM Edificações" in text or "JM Edificações —" in text:
                    continue

                if text.startswith("ⓘ"):
                    current_note = text.lstrip("ⓘ").strip(" -")
                    continue

                current_group = clean_group_title(text)
                current_note = None
                groups.append(
                    {
                        "sheet": worksheet.title,
                        "sheetSlug": slugify(worksheet.title),
                        "group": current_group,
                        "groupSlug": slugify(current_group),
                        "note": None,
                    }
                )
                continue

            if isinstance(a, str) and normalize_space(a) == "#":
                if groups and groups[-1]["sheet"] == worksheet.title and groups[-1]["group"] == current_group:
                    groups[-1]["note"] = current_note
                continue

            if not isinstance(a, (int, float)) or not b:
                continue

            description = normalize_space(b)
            unit: str | None
            reference_price: object | None
            jm_price: object | None

            if is_projects_sheet:
                unit = normalize_space(c) or None
                reference_price = None
                jm_price = parse_price(d)
            else:
                description, unit = extract_unit(description)
                reference_price = parse_price(c)
                jm_price = parse_price(d)

            item = build_item(
                sheet_name=worksheet.title,
                group_title=current_group,
                group_note=current_note,
                row_index=row_index,
                item_id=int(a),
                description=description,
                unit=unit,
                reference_price=reference_price,
                jm_price=jm_price,
            )
            items.append(item)

    numeric_prices = [item["jmPrice"] for item in items if isinstance(item["jmPrice"], (int, float))]
    consultation_count = sum(1 for item in items if item["priceMode"] == "consult")
    formula_count = sum(1 for item in items if item["priceMode"] == "formula")
    groups = sorted(groups, key=lambda entry: (str(entry["sheet"]), str(entry["group"])))
    areas = sorted({item["sheet"] for item in items})

    return {
        "meta": {
            "sourceFile": SOURCE_FILE.name,
            "sourceModifiedAt": datetime.fromtimestamp(SOURCE_FILE.stat().st_mtime).isoformat(),
            "generatedAt": datetime.now().isoformat(),
            "currency": "BRL",
            "summary": {
                "totalItems": len(items),
                "totalAreas": len(areas),
                "totalGroups": len(groups),
                "consultationItems": consultation_count,
                "formulaItems": formula_count,
                "pricedItems": len(numeric_prices),
                "minPrice": min(numeric_prices) if numeric_prices else None,
                "maxPrice": max(numeric_prices) if numeric_prices else None,
            },
        },
        "areas": areas,
        "groups": groups,
        "items": items,
    }


def main() -> None:
    payload = parse_workbook()
    serialized = json.dumps(payload, ensure_ascii=False, indent=2)
    OUTPUT_FILE.write_text(
        "// Arquivo gerado automaticamente a partir da planilha de precificação.\n"
        "// Edite o .xlsx e execute atualizar-precos.ps1 para sincronizar.\n"
        f"window.JM_PRICING_DATA = {serialized};\n",
        encoding="utf-8",
    )
    print(f"pricing-data.js atualizado com {payload['meta']['summary']['totalItems']} serviços.")


if __name__ == "__main__":
    main()
